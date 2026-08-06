"""
report_generator.py
Turns the JSON analysis returned by HomeworkChecker into three deliverables:

  1. generate_html_report()   -> a styled, shareable .html file
  2. generate_pdf_report()    -> a printable .pdf file (pure-python, no
                                  system dependencies like wkhtmltopdf)
  3. generate_excel_gradebook() -> a .xlsx gradebook across many students

Each function returns the Path it wrote to.
"""

from datetime import datetime
from pathlib import Path
from typing import List, Tuple

from jinja2 import Environment, FileSystemLoader, select_autoescape

from config import REPORTS_DIR, TEMPLATES_DIR, grade_for_score


# ---------------------------------------------------------------------------
# HTML report
# ---------------------------------------------------------------------------
def generate_html_report(result: dict, student_name: str = "Student", out_name: str = None) -> Path:
    env = Environment(
        loader=FileSystemLoader(str(TEMPLATES_DIR)),
        autoescape=select_autoescape(["html"]),
    )
    template = env.get_template("report.html")

    for q in result["questions"]:
        q["grade_label"], q["grade_color"] = grade_for_score(q["score"])
    overall_label, overall_color = grade_for_score(result["overall_score"])

    html = template.render(
        student_name=student_name,
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
        overall_score=result["overall_score"],
        grade=result.get("grade", overall_label),
        grade_color=overall_color,
        questions=result["questions"],
        summary=result.get("summary", {}),
    )

    out_name = out_name or f"report_{student_name.replace(' ', '_')}_{datetime.now():%Y%m%d%H%M%S}.html"
    out_path = REPORTS_DIR / out_name
    out_path.write_text(html, encoding="utf-8")
    return out_path


# ---------------------------------------------------------------------------
# PDF report (pure python via fpdf2 -- no system deps required)
# ---------------------------------------------------------------------------
def generate_pdf_report(result: dict, student_name: str = "Student", out_name: str = None) -> Path:
    from fpdf import FPDF

    pdf = FPDF(format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)

    # Helvetica (a core PDF font) only supports Latin-1 and crashes on
    # math symbols like √, ∫, ≠, ≤, π, etc. We ship a bundled Unicode
    # font (DejaVu Sans) so any symbol Gemini returns renders safely.
    font_dir = Path(__file__).resolve().parent / "static" / "fonts"
    regular_font = font_dir / "DejaVuSans.ttf"
    bold_font = font_dir / "DejaVuSans-Bold.ttf"
    use_unicode_font = regular_font.exists() and bold_font.exists()

    if use_unicode_font:
        pdf.add_font("DejaVu", "", str(regular_font))
        pdf.add_font("DejaVu", "B", str(bold_font))
        font_name = "DejaVu"
    else:
        # Fallback so the app still works if the font files are missing,
        # though unusual math symbols may then raise an encoding error.
        font_name = "Helvetica"

    pdf.add_page()

    def h1(text):
        pdf.set_font(font_name, "B", 16)
        pdf.multi_cell(0, 10, text, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1)

    def h2(text):
        pdf.set_font(font_name, "B", 12)
        pdf.multi_cell(0, 8, text, new_x="LMARGIN", new_y="NEXT")

    def body(text, size=10):
        pdf.set_font(font_name, "", size)
        pdf.multi_cell(0, 6, text, new_x="LMARGIN", new_y="NEXT")

    def bullet_list(items):
        pdf.set_font(font_name, "", 10)
        for item in items:
            pdf.multi_cell(0, 6, f"  -  {item}", new_x="LMARGIN", new_y="NEXT")

    overall_label, _ = grade_for_score(result["overall_score"])

    h1("Homework Grading Report")
    body(f"Student: {student_name}")
    body(f"Generated: {datetime.now():%Y-%m-%d %H:%M}")
    body(f"Overall Score: {result['overall_score']}/100  ({result.get('grade', overall_label)})")
    pdf.ln(4)

    for q in result["questions"]:
        pdf.set_draw_color(200, 200, 200)
        pdf.line(pdf.get_x(), pdf.get_y(), pdf.get_x() + 190, pdf.get_y())
        pdf.ln(3)

        h2(f"Question {q['question_number']}  -  Score: {q['score']}/100  ({q.get('difficulty', '')})")
        body(f"Question: {q['question']}")
        body(f"Student answer: {q['student_answer']}")
        body(f"Expected solution: {q['expected_solution']}")
        pdf.ln(1)
        body("Analysis:", size=10)
        body(q["analysis"])

        if q.get("mistakes"):
            pdf.ln(1)
            body("Mistakes identified:")
            bullet_list(q["mistakes"])

        if q.get("error_categories"):
            body("Error categories: " + ", ".join(q["error_categories"]))

        if q.get("correct_answer"):
            pdf.ln(1)
            body(f"Correct answer: {q['correct_answer']}")

        if q.get("suggestions"):
            pdf.ln(1)
            body("Suggestions:")
            bullet_list(q["suggestions"])

        pdf.ln(4)

    summary = result.get("summary", {})
    pdf.add_page()
    h1("Final Summary")
    for label, key in [
        ("Strengths", "strengths"),
        ("Weaknesses", "weaknesses"),
        ("Topics to Review", "topics_to_review"),
        ("Study Recommendations", "recommendations"),
    ]:
        items = summary.get(key) or []
        if items:
            h2(label)
            bullet_list(items)
            pdf.ln(2)

    out_name = out_name or f"report_{student_name.replace(' ', '_')}_{datetime.now():%Y%m%d%H%M%S}.pdf"
    out_path = REPORTS_DIR / out_name
    pdf.output(str(out_path))
    return out_path


# ---------------------------------------------------------------------------
# Excel gradebook (across one or many students/submissions)
# ---------------------------------------------------------------------------
def generate_excel_gradebook(
    records: List[Tuple[str, dict]],
    out_name: str = None,
) -> Path:
    """
    records: list of (student_name, result_dict) tuples.
    Produces a workbook with:
      - "Gradebook" sheet: one row per student with overall score/grade
      - "Question Detail" sheet: one row per question across all students
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # --- Sheet 1: Gradebook overview ---
    ws1 = wb.active
    ws1.title = "Gradebook"
    headers1 = ["Student", "Overall Score", "Grade", "Questions Detected", "Generated At"]
    ws1.append(headers1)
    for col in range(1, len(headers1) + 1):
        cell = ws1.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for student_name, result in records:
        label, _ = grade_for_score(result["overall_score"])
        ws1.append([
            student_name,
            result["overall_score"],
            result.get("grade", label),
            len(result.get("questions", [])),
            datetime.now().strftime("%Y-%m-%d %H:%M"),
        ])

    for i, width in enumerate([24, 14, 18, 18, 18], start=1):
        ws1.column_dimensions[get_column_letter(i)].width = width

    # --- Sheet 2: Per-question detail ---
    ws2 = wb.create_sheet("Question Detail")
    headers2 = [
        "Student", "Q#", "Score", "Difficulty", "Error Categories",
        "Mistakes", "Correct Answer",
    ]
    ws2.append(headers2)
    for col in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for student_name, result in records:
        for q in result.get("questions", []):
            ws2.append([
                student_name,
                q.get("question_number"),
                q.get("score"),
                q.get("difficulty"),
                ", ".join(q.get("error_categories", [])),
                "; ".join(q.get("mistakes", [])),
                q.get("correct_answer"),
            ])

    for i, width in enumerate([24, 6, 8, 12, 24, 40, 30], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = width
    ws2.freeze_panes = "A2"
    ws1.freeze_panes = "A2"

    out_name = out_name or f"gradebook_{datetime.now():%Y%m%d%H%M%S}.xlsx"
    out_path = REPORTS_DIR / out_name
    wb.save(str(out_path))
    return out_path
